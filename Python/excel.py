import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from pandas import ExcelWriter
from pandas.io.formats.style import Styler

from clan import Clan


class Excel(Clan):
    def __init__(self, clan_tag: str, url: str, headers: str):
        super().__init__(url, headers)
        self.clan_tag: str = clan_tag
        self.average_donation: int = sum([d['donations'] for d in self.get_clan_info()]) / len(self.get_clan_info())

    def create_clan_overview_excel_file(self) -> None:
        maximum_colormap_value = (max(self.get_donations_of_member()) / self.get_number_of_members()) * 10
        df: Styler = pd.DataFrame(self.format_role_name()).sort_values(by=['Donations'],
                                                                       ascending=False).style.background_gradient(
            subset=["Donations"],
            cmap="RdYlGn",
            vmin=0,
            vmax=maximum_colormap_value,
            low=0, high=1,
            axis=1)

        writer: ExcelWriter = pd.ExcelWriter(f'Clan Overview - #{self.clan_tag}.xlsx', engine='xlsxwriter')
        df.to_excel(writer, sheet_name=datetime.datetime.now().strftime('%d-%m-%Y'), index=False)

        writer.save()

        # Load the Excel workbook
        workbook = load_workbook(f'Clan Overview - #{self.clan_tag}.xlsx')
        # Select the worksheet you want to work with
        worksheet = workbook[datetime.datetime.now().strftime('%d-%m-%Y')]

        # Set the row height of all rows in the worksheet to 30
        for row in worksheet.iter_rows():
            worksheet.row_dimensions[row[0].row].height = 20

        # Set the column width based on the maximum length of cell values in each column
        for column in worksheet.columns:
            max_length = 0
            for cell in column:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            worksheet.column_dimensions[
                column[0].column_letter].width = max_length * 1.3  # adjust the multiplier as needed

        # Set the alignment of all cells in the worksheet to center
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Save the changes to the workbook
        workbook.save(f'Clan Overview - #{self.clan_tag}.xlsx')
